import openpyxl
import datetime
from openpyxl.chart import (
    Reference,
    BarChart
)


book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"
sheet1.cell(row=1, column=1).value = 'Title'
sheet1.cell(row=2, column=2).value = '에벵벵'
sheet1.cell(row=4, column=4).value = '바꾸어라'
sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
sheet2['A1'] = datetime.datetime.now()
sheet2['A2'] = datetime.date.today()
# 저장하기
book.save("./output.xlsx")

# import openpyxl
# # insert image
# imgFile = './web/aaa.png'
# img = openpyxl.drawing.image.Image(imgFile)
# sheet1.add_image(img, 'B5')
# # resize image
# from PIL import Image
# img2 = Image.open(imgFile)
# new_img = img2.resize((100, 100))
# new_img.save('new.png')
# img3 = openpyxl.drawing.image.Image(new.png)
# sheet1.add_image(img3, 'A5')

rows = [
    ['김일수', 11],
    ['김이수', 22],
    ['김삼수', 33],
    ['김사수', 15],
    ['김오수', 11],
]

for row in rows:
    sheet1.append(row)
datax = Reference(sheet1, min_col=2, 
		min_row=1, max_col=2, max_row=5)
categs = Reference(sheet1, min_col=1,
				 min_row=1, max_row=5)

chart = BarChart()
chart.add_data(data=datax)
chart.set_categories(categs)

chart.legend = None  # 범례
chart.varyColors = True
chart.title = "차트 타이틀"

sheet1.add_chart(chart, "A8")